# file: pipeline.py
import argparse
import logging
from typing import Dict, Optional, List, Any

import pandas as pd
from pydantic import BaseModel, ValidationError, Field

try:
    from transformers import AutoTokenizer, AutoModelForSeq2SeqLM
    import torch
    HAS_TRANSFORMERS = True
except ImportError:
    HAS_TRANSFORMERS = False

from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)


# -----------------------------
# 1. Pydantic schemas
# -----------------------------

class AgentsRow(BaseModel):
    agent_name: str = Field(min_length=1, max_length=256)
    business_name: Optional[str] = Field(default=None, max_length=256)


class PipelineConfig(BaseModel):
    sort_by: Optional[str] = None
    drop_columns: List[str] = []
    money_columns: List[str] = []
    model_name: str = "google/flan-t5-small"
    max_prompt_tokens: int = 512
    start_row: int = 10
    gap_rows: int = 4


# -----------------------------
# 2. Input & Validation
# -----------------------------

def load_file(path: str) -> pd.DataFrame:
    logger.info("Loading file: %s", path)

    if path.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        df = pd.read_excel(path)
    elif path.lower().endswith(".csv"):
        df = pd.read_csv(path)
    else:
        raise ValueError(f"Unsupported file type: {path}")

    logger.info("Loaded %d rows, %d columns from %s", len(df), len(df.columns), path)
    return df


def validate_agents_df(df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ["agent_name", "business_name"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"agents file missing required column: {col}")

    valid_rows = []
    for idx, row in df.iterrows():
        try:
            validated = AgentsRow(
                agent_name=str(row["agent_name"]).strip(),
                business_name=None if pd.isna(row["business_name"]) else str(row["business_name"]).strip()
            )
            valid_rows.append(validated.model_dump())
        except ValidationError as e:
            logger.error("Invalid agents row at index %d: %s", idx, e)
            continue

    logger.info("Validated %d agent rows (out of %d)", len(valid_rows), len(df))
    return pd.DataFrame(valid_rows)


def validate_data_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        logger.warning("export data is empty")
    for col in df.columns:
        if len(col) == 0 or len(col) > 128:
            raise ValueError(f"Invalid column name length for: {col!r}")
    return df


# -----------------------------
# 3. Transformation
# -----------------------------

def build_agent_map(agents_df: pd.DataFrame) -> Dict[str, Optional[str]]:
    agent_map: Dict[str, Optional[str]] = {}
    for _, row in agents_df.iterrows():
        name = row["agent_name"].strip()
        business = row["business_name"]
        agent_map[name] = business if isinstance(business, str) and business else None
    logger.info("Built agent_map with %d entries", len(agent_map))
    return agent_map


def normalize_data_df(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        if pd.api.types.is_string_dtype(df[col]):
            df[col] = df[col].astype(str).str.strip()
    return df


# -----------------------------
# 4. Auto-detect status column
# -----------------------------

def detect_status_column(df: pd.DataFrame) -> str:
    """
    Auto-detect the column containing A/C/P status codes.
    Must contain at least one A/C/P value.
    Must NOT contain values outside A/C/P (ignoring blanks).
    """
    candidates = []

    for col in df.columns:
        series = df[col].dropna().astype(str).str.strip()

        # Skip columns that are entirely empty
        if series.empty:
            continue

        # Values that are not blank
        non_blank = series[series != ""]

        # Skip if no real values
        if non_blank.empty:
            continue

        # Check if all non-blank values are A/C/P
        if all(v in {"A", "C", "P"} for v in non_blank.unique()):
            candidates.append(col)

    if len(candidates) == 1:
        logger.info("Auto-detected status column: %s", candidates[0])
        return candidates[0]

    if len(candidates) == 0:
        raise ValueError(
            "Could not auto-detect status column. "
            "No column contains A/C/P values."
        )

    raise ValueError(
        f"Multiple possible status columns detected: {candidates}. "
        "Please clean your data or specify manually."
    )



# -----------------------------
# 5. Representation
# -----------------------------

def build_prompt_for_row(row: pd.Series, agent_map: Dict[str, Optional[str]], config: PipelineConfig) -> str:
    row_dict = row.to_dict()
    prompt = (
        "You are a data cleaning assistant.\n"
        "Given a table row and a mapping from agent names to business names, "
        "perform the following:\n"
        "1. Replace any agent name in the row with the corresponding business name.\n"
        "2. If no business name exists, keep the agent name but wrap it in **bold**.\n"
        "3. Keep all other fields unchanged.\n"
        "4. Return the row as a JSON object with the same keys.\n\n"
        f"Agent mapping: {agent_map}\n"
        f"Row: {row_dict}\n"
        "Output JSON:"
    )
    return prompt


# -----------------------------
# 6. Model Inference (FLAN-T5)
# -----------------------------

class FlanT5Wrapper:
    def __init__(self, model_name: str):
        if not HAS_TRANSFORMERS:
            raise RuntimeError("transformers not installed; cannot use FLAN-T5")
        logger.info("Loading model: %s", model_name)
        self.tokenizer = AutoTokenizer.from_pretrained(model_name)
        self.model = AutoModelForSeq2SeqLM.from_pretrained(model_name)
        self.device = "cuda" if torch.cuda.is_available() else "cpu"
        self.model.to(self.device)
        logger.info("Model loaded on device: %s", self.device)

    def generate(self, prompts: List[str], max_new_tokens: int = 256) -> List[str]:
        outputs: List[str] = []
        for prompt in prompts:
            inputs = self.tokenizer(prompt, return_tensors="pt", truncation=True)
            inputs = {k: v.to(self.device) for k, v in inputs.items()}
            with torch.no_grad():
                generated = self.model.generate(**inputs, max_new_tokens=max_new_tokens)
            text = self.tokenizer.decode(generated[0], skip_special_tokens=True)
            outputs.append(text)
        return outputs


# -----------------------------
# 7. Postprocessing
# -----------------------------

def deterministic_agent_replace(row: pd.Series, agent_map: Dict[str, Optional[str]]) -> pd.Series:
    def replace_text(text: Any) -> Any:
        if not isinstance(text, str):
            return text
        for agent, business in agent_map.items():
            if agent in text:
                replacement = f"**{business}**" if business else f"**{agent}**"
                text = text.replace(agent, replacement)
        return text

    return row.apply(replace_text)


def parse_model_output_to_row(model_output: str, original_row: pd.Series, agent_map: Dict[str, Optional[str]]) -> pd.Series:
    import json
    try:
        parsed = json.loads(model_output)
        for key in original_row.index:
            if key not in parsed:
                parsed[key] = original_row[key]
        return pd.Series(parsed)[original_row.index]
    except Exception:
        return deterministic_agent_replace(original_row, agent_map)


def format_numeric_columns(df: pd.DataFrame, config: PipelineConfig) -> pd.DataFrame:
    for col in config.money_columns:
        if col not in df.columns:
            continue

        def fmt(x):
            if pd.isna(x):
                return x
            try:
                val = float(x)
                return f"${val:,.2f}"
            except Exception:
                return x

        df[col] = df[col].apply(fmt)
    return df


def apply_sort_and_drop(df: pd.DataFrame, config: PipelineConfig) -> pd.DataFrame:
    if config.sort_by and config.sort_by in df.columns:
        df = df.sort_values(by=config.sort_by)
    for col in config.drop_columns:
        if col in df.columns:
            df = df.drop(columns=[col])
    return df


# -----------------------------
# 8. Excel template integration
# -----------------------------

def split_by_status(df: pd.DataFrame, status_column: str) -> Dict[str, pd.DataFrame]:
    sections = {
        "A": df[df[status_column] == "A"],
        "C": df[df[status_column] == "C"],
        "P": df[df[status_column] == "P"],
    }
    return sections


def write_sections_into_template(
    worksheet,
    sections: Dict[str, pd.DataFrame],
    start_row: int,
    gap_rows: int
) -> None:
    """
    Writes A, C, P sections directly under the last filled template row.
    Removes unused placeholder rows and preserves template formatting.
    """

    def last_filled_row(ws):
        """Return the last row that contains ANY non-empty cell."""
        max_row = ws.max_row
        for r in range(max_row, 0, -1):
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=r, column=c).value not in (None, ""):
                    return r
        return 1

    # Start writing at the first template data row
    current_row = max(start_row, last_filled_row(worksheet) + 1)

    for key in ["A", "C", "P"]:
        section = sections.get(key)
        if section is None or section.empty:
            continue

        # Write this section
        for _, row in section.iterrows():
            for col_idx, value in enumerate(row.values, start=1):
                worksheet.cell(row=current_row, column=col_idx, value=value)
            current_row += 1

        # Optional gap between sections
        current_row += gap_rows

    # ---- CLEANUP: Remove leftover blank template rows below the last written row ----
    final_last = last_filled_row(worksheet)
    max_row = worksheet.max_row

    for r in range(final_last + 1, max_row + 1):
        for c in range(1, worksheet.max_column + 1):
            worksheet.cell(row=r, column=c, value=None)



# -----------------------------
# 9. Full pipeline orchestration
# -----------------------------

def run_pipeline(
    template_path: str,
    export_path: str,
    agents_path: str,
    output_path: str,
    config: PipelineConfig
) -> None:
    # Load template workbook
    logger.info("Loading template workbook: %s", template_path)
    wb = load_workbook(template_path)
    ws = wb.active

    # Load inputs
    agents_df_raw = load_file(agents_path)
    export_df_raw = load_file(export_path)

    # Validate
    agents_df = validate_agents_df(agents_df_raw)
    data_df = validate_data_df(export_df_raw)

    # Normalize
    agent_map = build_agent_map(agents_df)
    data_df = normalize_data_df(data_df)

    # Auto-detect status column
    status_col = detect_status_column(data_df)

    # Model or deterministic replacement
    use_model = HAS_TRANSFORMERS
    model = None
    processed_rows = []

    if use_model:
        try:
            model = FlanT5Wrapper(config.model_name)
            prompts = [build_prompt_for_row(row, agent_map, config) for _, row in data_df.iterrows()]
            outputs = model.generate(prompts)
            for (idx, row), out in zip(data_df.iterrows(), outputs):
                processed_rows.append(parse_model_output_to_row(out, row, agent_map))
        except Exception as e:
            logger.error("Model failed, falling back to deterministic mode: %s", e)
            use_model = False

    if not use_model:
        for _, row in data_df.iterrows():
            processed_rows.append(deterministic_agent_replace(row, agent_map))

    processed_df = pd.DataFrame(processed_rows)

    # Postprocessing
    processed_df = format_numeric_columns(processed_df, config)
    processed_df = apply_sort_and_drop(processed_df, config)

    # Split into A/C/P
    sections = split_by_status(processed_df, status_col)

    # Write into template
    write_sections_into_template(
        worksheet=ws,
        sections=sections,
        start_row=config.start_row,
        gap_rows=config.gap_rows,
    )

    # Save output
    wb.save(output_path)
    logger.info("Wrote output Excel file to %s", output_path)


# -----------------------------
# 10. CLI entrypoint
# -----------------------------

def main():
    parser = argparse.ArgumentParser(description="FLAN-T5 Excel Name Normalization Pipeline")
    parser.add_argument("--template", required=True, help="Path to blank template .xlsx")
    parser.add_argument("--export", required=True, help="Path to exported data .xltx/.xlsx/.csv")
    parser.add_argument("--agents", required=True, help="Path to agents .xlsx/.csv")
    parser.add_argument("--output", required=True, help="Path to output .xlsx")
    parser.add_argument("--sort-by", default=None)
    parser.add_argument("--drop-columns", nargs="*", default=[])
    parser.add_argument("--money-columns", nargs="*", default=[])
    parser.add_argument("--model-name", default="google/flan-t5-small")
    parser.add_argument("--start-row", type=int, default=10)
    parser.add_argument("--gap-rows", type=int, default=4)
    args = parser.parse_args()

    config = PipelineConfig(
        sort_by=args.sort_by,
        drop_columns=args.drop_columns,
        money_columns=args.money_columns,
        model_name=args.model_name,
        start_row=args.start_row,
        gap_rows=args.gap_rows,
    )

    run_pipeline(
        template_path=args.template,
        export_path=args.export,
        agents_path=args.agents,
        output_path=args.output,
        config=config,
    )


if __name__ == "__main__":
    main()
